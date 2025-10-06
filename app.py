import io
import os
import re
import zipfile
from datetime import datetime

import streamlit as st
import pandas as pd

from PIL import Image
import qrcode, qrcode.image.svg

# ReportLab (PDF)
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Arabic shaping/bidi
import arabic_reshaper
from bidi.algorithm import get_display

# Excel template
from openpyxl import Workbook

# -------------------------------------------------
# GLOBAL CONFIG
# -------------------------------------------------

APP_TITLE = "Internal Hub"

# Minimal table styling
st.markdown("""
<style>
[data-testid="stDataFrame"] table {border: 1px solid #e8ecf2; border-radius: 8px;}
[data-testid="stDataFrame"] th {background-color: #f5f8fc !important; color: #254489 !important; font-weight: 600 !important;}
</style>
""", unsafe_allow_html=True)

# Fonts (must be valid TTFs with Arabic glyphs)
FONT_BOLD_PATH = "assets/Fonts/PingAR+LT-Bold.ttf"
FONT_REGULAR_PATH = "assets/Fonts/PingAR+LT-Regular.ttf"

# Backgrounds for signatures (case-sensitive paths!)
SIG_BG_EN = "assets/images/eng-bg.png"
SIG_BG_AR = "assets/images/arbc-bg.png"

# Business card faces (you used capital 'Images' here)
CARD_FRONT = "assets/Images/front.png"
CARD_BACK  = "assets/Images/back.png"

# Brand color
COLOR_HEX = "#254489"
COLOR_RGB_REPORTLAB = (0.145, 0.266, 0.537)  # #254489

# Register fonts for ReportLab (used by PDFs)
pdfmetrics.registerFont(TTFont("PingBold",    FONT_BOLD_PATH))
pdfmetrics.registerFont(TTFont("PingRegular", FONT_REGULAR_PATH))

# -------------------------------------------------
# DEPARTMENTS & ROLES
# -------------------------------------------------

DEPARTMENTS = {
    "PMO": ["Head of Strategy","Head of PMO","IT Program Manager","Business Analyst","Board Member & CEO"],
    "IT": ["Head of IT","Senior Manager of Technology Operation & Applications","Senior Technical & Network Engineer",
           "Senior DevOps Engineer","Senior Analyst","IT Support Administrator","Outsourced"],
    "Engineering & Data": ["Head of Engineering","Head of Data","Senior Software Engineer","Staff IOS Engineer",
                           "Data Engineer","Associate Data Engineer","Data Analyst","Senior Analyst","Analyst","Outsourced"],
    "Product": ["Principal","Lead","Senior"],
    "Cybersecurity": ["CISO","Cybersecurity Engineer","Cybersecurity Analyst","Threat Intelligence Analyst","Outsourced"],
    "Risk": ["Head of Risk","AVP - Operations Manager","AVP","Credit Manager","Senior Associate","Senior Operations Analyst",
             "Senior Analyst","Credit Underwriting Senior Analyst","Credit Underwriting Analyst","Credit Administration Analyst",
             "Operations Analyst","Associate","Analyst"],
    "Legal": ["Senior Corporate Government Analyst","Collection Manager","Senior Legal Analyst","Senior Associate","Team Leader",
              "Legal Analyst","Senior Analyst","Collection Officer","Associate","Analyst","Outsourced"],
    "Customer Care": ["Head of Customer Care","Senior Customer Care Specialist","Customer Care Specialist"],
    "Compliance": ["VP - Head of Compliance","Senior AML Officer","AML & Compliance Officer","Compliance Specialist"],
    "Digital Lending": ["Commercial Director","Digital Marketing Manager","Customer Success Manager","Customer Success Specialist"],
    "Direct Lending": ["Head of Sales","Senior Key Accounts Manager","Key Accounts Manager","Sales Manager","Senior Team Leader",
                       "Team Leader","Growth Manager","Growth Consultant II","Growth Consultant I","Growth Associate",
                       "Sales Operations","Business Development Representative","Quality Assurance Specialist"],
    "Marketing": ["Senior Communication Specialist","Marketing Specialist","Social Media Specialist","Marketing Executive"],
    "Finance": ["VP - Head of Finance","VP - Finance","Senior Finance Analyst","Senior Financial Associate","Financial Analyst","Financial Associate"]
}

ROLES_MAP = {
    "Head of Strategy":"ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑÿßÿ≥ÿ™ÿ±ÿßÿ™Ÿäÿ¨Ÿäÿ©","Head of PMO":"ÿ±ÿ¶Ÿäÿ≥ ŸÖŸÉÿ™ÿ® ÿ•ÿØÿßÿ±ÿ© ÿßŸÑŸÖÿ¥ÿßÿ±Ÿäÿπ","IT Program Manager":"ŸÖÿØŸäÿ± ÿ®ÿ±ŸÜÿßŸÖÿ¨ ÿ™ŸÇŸÜŸäÿ© ÿßŸÑŸÖÿπŸÑŸàŸÖÿßÿ™",
    "Business Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ£ÿπŸÖÿßŸÑ","Head of IT":"ÿ±ÿ¶Ÿäÿ≥ ÿ™ŸÇŸÜŸäÿ© ÿßŸÑŸÖÿπŸÑŸàŸÖÿßÿ™","Board Member & CEO":"ÿπÿ∂Ÿà ŸÖÿ¨ŸÑÿ≥ ÿßŸÑÿ•ÿØÿßÿ±ÿ© ŸàÿßŸÑÿ±ÿ¶Ÿäÿ≥ ÿßŸÑÿ™ŸÜŸÅŸäÿ∞Ÿä",
    "Senior Manager of Technology Operation & Applications":"ŸÖÿØŸäÿ± ÿ£ŸàŸÑ ŸÑÿπŸÖŸÑŸäÿßÿ™ ÿßŸÑÿ™ŸÇŸÜŸäÿ© ŸàÿßŸÑÿ™ÿ∑ÿ®ŸäŸÇÿßÿ™","Senior Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ£ŸàŸÑ",
    "Outsourced":"ŸÖŸàÿ∏ŸÅ ÿÆÿßÿ±ÿ¨Ÿä","Senior Technical & Network Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿ™ŸÇŸÜŸä Ÿàÿ¥ÿ®ŸÉÿßÿ™ ÿ£ŸàŸÑ","IT Support Administrator":"ŸÖÿ≥ÿ§ŸàŸÑ ÿØÿπŸÖ ÿ™ŸÇŸÜŸäÿ© ÿßŸÑŸÖÿπŸÑŸàŸÖÿßÿ™",
    "Senior DevOps Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿØŸäŸÅ ÿ£Ÿàÿ®ÿ≥ ÿ£ŸàŸÑ","Head of Engineering":"ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑŸáŸÜÿØÿ≥ÿ©","Senior Software Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿ®ÿ±ŸÖÿ¨Ÿäÿßÿ™ ÿ£ŸàŸÑ",
    "Staff IOS Engineer":"ŸÖŸáŸÜÿØÿ≥ ŸÜÿ∏ŸÖ iOS","Head of Data":"ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑÿ®ŸäÿßŸÜÿßÿ™","Data Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿ®ŸäÿßŸÜÿßÿ™","Analyst":"ŸÖÿ≠ŸÑŸÑ",
    "Associate Data Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿ®ŸäÿßŸÜÿßÿ™ ŸÖÿ¥ÿßÿ±ŸÉ","Data Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ®ŸäÿßŸÜÿßÿ™","Principal":"ÿ±ÿ¶Ÿäÿ≥ ŸÇÿ≥ŸÖ","Lead":"ŸÇÿßÿ¶ÿØ ŸÅÿ±ŸäŸÇ","Senior":"ŸÉÿ®Ÿäÿ±",
    "CISO":"ÿ±ÿ¶Ÿäÿ≥ ÿ£ŸÖŸÜ ÿßŸÑŸÖÿπŸÑŸàŸÖÿßÿ™","Cybersecurity Engineer":"ŸÖŸáŸÜÿØÿ≥ ÿ£ŸÖŸÜ ÿ≥Ÿäÿ®ÿ±ÿßŸÜŸä","Cybersecurity Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ£ŸÖŸÜ ÿ≥Ÿäÿ®ÿ±ÿßŸÜŸä",
    "Threat Intelligence Analyst":"ŸÖÿ≠ŸÑŸÑ ÿßÿ≥ÿ™ÿÆÿ®ÿßÿ±ÿßÿ™ ÿßŸÑÿ™ŸáÿØŸäÿØÿßÿ™","Head of Risk":"ÿ±ÿ¶Ÿäÿ≥ ÿ•ÿØÿßÿ±ÿ© ÿßŸÑŸÖÿÆÿßÿ∑ÿ±","Risk Officer":"ŸÖÿ≥ÿ§ŸàŸÑ ŸÖÿÆÿßÿ∑ÿ±","AVP":"ŸÖÿ≥ÿßÿπÿØ ŸÜÿßÿ¶ÿ® ÿßŸÑÿ±ÿ¶Ÿäÿ≥",
    "Credit Manager":"ŸÖÿØŸäÿ± ÿßŸÑÿßÿ¶ÿ™ŸÖÿßŸÜ","Credit Underwriting Senior Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ£ŸàŸÑ ŸÑŸÑÿßŸÉÿ™ÿ™ÿßÿ® ÿßŸÑÿßÿ¶ÿ™ŸÖÿßŸÜŸä",
    "Credit Underwriting Analyst":"ŸÖÿ≠ŸÑŸÑ ÿßŸÉÿ™ÿ™ÿßÿ® ÿßÿ¶ÿ™ŸÖÿßŸÜŸä","AVP - Operations Manager":"ŸÖÿ≥ÿßÿπÿØ ŸÜÿßÿ¶ÿ® ÿßŸÑÿ±ÿ¶Ÿäÿ≥ - ŸÖÿØŸäÿ± ÿßŸÑÿπŸÖŸÑŸäÿßÿ™",
    "Senior Associate":"ŸÖÿ≥ÿßÿπÿØ ÿ£ŸàŸÑ","Associate":"ŸÖÿ≥ÿßÿπÿØ","Senior Operations Analyst":"ŸÖÿ≠ŸÑŸÑ ÿπŸÖŸÑŸäÿßÿ™ ÿ£ŸàŸÑ","Operations Analyst":"ŸÖÿ≠ŸÑŸÑ ÿπŸÖŸÑŸäÿßÿ™",
    "Credit Administration Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ•ÿØÿßÿ±ÿ© ÿßŸÑÿßÿ¶ÿ™ŸÖÿßŸÜ","Senior Legal Analyst":"ŸÖÿ≠ŸÑŸÑ ŸÇÿßŸÜŸàŸÜŸä ÿ£ŸàŸÑ","Legal Analyst":"ŸÖÿ≠ŸÑŸÑ ŸÇÿßŸÜŸàŸÜŸä",
    "Team Leader":"ŸÇÿßÿ¶ÿØ ŸÅÿ±ŸäŸÇ","Collection Manager":"ŸÖÿØŸäÿ± ÿßŸÑÿ™ÿ≠ÿµŸäŸÑ","Senior Collection Officer":"ŸÖÿ≥ÿ§ŸàŸÑ ÿ™ÿ≠ÿµŸäŸÑ ÿ£ŸàŸÑ","Collection Officer":"ŸÖÿ≥ÿ§ŸàŸÑ ÿ™ÿ≠ÿµŸäŸÑ",
    "Senior Corporate Government Analyst":"ŸÖÿ≠ŸÑŸÑ ÿ≠ŸàŸÉŸÖÿ© ŸÖÿ§ÿ≥ÿ≥Ÿäÿ© ÿ£ŸàŸÑ","Head of Customer Care":"ÿ±ÿ¶Ÿäÿ≥ ÿÆÿØŸÖÿ© ÿßŸÑÿπŸÖŸÑÿßÿ°",
    "Senior Customer Care Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿÆÿØŸÖÿ© ÿßŸÑÿπŸÖŸÑÿßÿ° ÿ£ŸàŸÑ","Customer Care Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿÆÿØŸÖÿ© ÿßŸÑÿπŸÖŸÑÿßÿ°",
    "VP - Head of Compliance":"ŸÜÿßÿ¶ÿ® ÿßŸÑÿ±ÿ¶Ÿäÿ≥ - ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑÿßŸÖÿ™ÿ´ÿßŸÑ","Senior AML Officer":"ŸÖÿ≥ÿ§ŸàŸÑ ÿ£ŸàŸÑ ŸÑŸÖŸÉÿßŸÅÿ≠ÿ© ÿ∫ÿ≥ŸÑ ÿßŸÑÿ£ŸÖŸàÿßŸÑ",
    "AML & Compliance Officer":"ŸÖÿ≥ÿ§ŸàŸÑ ÿßŸÖÿ™ÿ´ÿßŸÑ ŸàŸÖŸÉÿßŸÅÿ≠ÿ© ÿ∫ÿ≥ŸÑ ÿßŸÑÿ£ŸÖŸàÿßŸÑ","Compliance Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿßŸÖÿ™ÿ´ÿßŸÑ",
    "Commercial Director":"ÿßŸÑŸÖÿØŸäÿ± ÿßŸÑÿ™ÿ¨ÿßÿ±Ÿä","Digital Marketing Manager":"ŸÖÿØŸäÿ± ÿßŸÑÿ™ÿ≥ŸàŸäŸÇ ÿßŸÑÿ±ŸÇŸÖŸä","Customer Success Manager":"ŸÖÿØŸäÿ± ŸÜÿ¨ÿßÿ≠ ÿßŸÑÿπŸÖŸÑÿßÿ°",
    "Customer Success Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ŸÜÿ¨ÿßÿ≠ ÿßŸÑÿπŸÖŸÑÿßÿ°","Head of Sales":"ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑŸÖÿ®Ÿäÿπÿßÿ™","Sales Manager":"ŸÖÿØŸäÿ± ŸÖÿ®Ÿäÿπÿßÿ™",
    "Senior Team Leader":"ŸÇÿßÿ¶ÿØ ŸÅÿ±ŸäŸÇ ÿ£ŸàŸÑ","Growth Manager":"ŸÖÿØŸäÿ± ÿßŸÑŸÜŸÖŸà","Growth Consultant II":"ŸÖÿ≥ÿ™ÿ¥ÿßÿ± ŸÜŸÖŸà II","Growth Consultant I":"ŸÖÿ≥ÿ™ÿ¥ÿßÿ± ŸÜŸÖŸà I",
    "Growth Associate":"ŸÖÿ≥ÿßÿπÿØ ŸÜŸÖŸà","Key Accounts Manager":"ŸÖÿØŸäÿ± ÿßŸÑÿ≠ÿ≥ÿßÿ®ÿßÿ™ ÿßŸÑÿ±ÿ¶Ÿäÿ≥Ÿäÿ©","Senior Key Accounts Manager":"ŸÖÿØŸäÿ± ÿßŸÑÿ≠ÿ≥ÿßÿ®ÿßÿ™ ÿßŸÑÿ±ÿ¶Ÿäÿ≥Ÿäÿ© ÿ£ŸàŸÑ",
    "Sales Operations":"ÿπŸÖŸÑŸäÿßÿ™ ÿßŸÑŸÖÿ®Ÿäÿπÿßÿ™","Business Development Representative":"ŸÖŸÖÿ´ŸÑ ÿ™ÿ∑ŸàŸäÿ± ÿßŸÑÿ£ÿπŸÖÿßŸÑ","Quality Assurance Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿ∂ŸÖÿßŸÜ ÿßŸÑÿ¨ŸàÿØÿ©",
    "Senior Communication Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿßÿ™ÿµÿßŸÑÿßÿ™ ÿ£ŸàŸÑ","Marketing Executive":"ÿ™ŸÜŸÅŸäÿ∞Ÿä ÿ™ÿ≥ŸàŸäŸÇ","Social Media Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä Ÿàÿ≥ÿßÿ¶ŸÑ ÿßŸÑÿ™ŸàÿßÿµŸÑ ÿßŸÑÿßÿ¨ÿ™ŸÖÿßÿπŸä",
    "Marketing Specialist":"ÿ£ÿÆÿµÿßÿ¶Ÿä ÿ™ÿ≥ŸàŸäŸÇ","VP - Head of Finance":"ŸÜÿßÿ¶ÿ® ÿßŸÑÿ±ÿ¶Ÿäÿ≥ - ÿ±ÿ¶Ÿäÿ≥ ÿßŸÑŸÖÿßŸÑŸäÿ©","VP - Finance":"ŸÜÿßÿ¶ÿ® ÿßŸÑÿ±ÿ¶Ÿäÿ≥ - ÿßŸÑŸÖÿßŸÑŸäÿ©",
    "Senior Financial Associate":"ŸÖÿ≥ÿßÿπÿØ ŸÖÿßŸÑŸä ÿ£ŸàŸÑ","Financial Associate":"ŸÖÿ≥ÿßÿπÿØ ŸÖÿßŸÑŸä","Senior Finance Analyst":"ŸÖÿ≠ŸÑŸÑ ŸÖÿßŸÑŸä ÿ£ŸàŸÑ","Financial Analyst":"ŸÖÿ≠ŸÑŸÑ ŸÖÿßŸÑŸä"
}

# -------------------------------------------------
# HELPERS
# -------------------------------------------------

def arabic_ready(text: str) -> str:
    return get_display(arabic_reshaper.reshape(text)) if text else ""

def normalize_saudi_mobile(mobile: str) -> tuple[str, bool]:
    if not mobile: return ("", False)
    clean = re.sub(r"\D", "", str(mobile))
    if clean.startswith("00966") and len(clean) == 14: clean = clean[2:]
    if clean.startswith("9665") and len(clean) == 12: return (f"+{clean}", True)
    if clean.startswith("05")   and len(clean) == 10: return (f"+966{clean[1:]}", True)
    if clean.startswith("5")    and len(clean) == 9:  return (f"+966{clean}", True)
    if clean.startswith("966")  and len(clean) == 12: return (f"+{clean}", True)
    if clean.startswith("+966") and len(clean) == 13: return (clean, True)
    return (mobile, False)

def normalize_email(email: str) -> tuple[str, bool]:
    if not email: return ("", False)
    e = email.strip().lower()
    return (e, "@" in e and e.count("@") == 1)

def normalize_name(name: str) -> str:
    return name.strip().capitalize() if name else ""

def vcard_from_person(p: dict) -> str:
    first = p.get("First_Name","") or ""
    last  = p.get("Last_Name","")  or ""
    org   = p.get("Company","")    or ""
    title = p.get("Role","")       or ""
    tel   = p.get("Mobile","")     or ""
    email = p.get("Email","")      or ""
    url   = p.get("Website","")    or ""
    loc   = p.get("Location","")   or ""
    notes = p.get("Notes","")      or ""
    return f"""BEGIN:VCARD
VERSION:3.0
N:{last};{first}
FN:{first} {last}
ORG:{org}
TITLE:{title}
TEL;TYPE=CELL:{tel}
EMAIL:{email}
URL:{url}
ADR;TYPE=WORK:{loc}
NOTE:{notes}
END:VCARD"""

def make_qr_png_bytes(data: str, fill_color="#254489") -> bytes:
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(data or "")
    qr.make(fit=True)
    img = qr.make_image(fill_color=fill_color, back_color="white").convert("RGBA")
    pixels = img.getdata()
    img.putdata([(r,g,b,0) if r>240 and g>240 and b>240 else (r,g,b,a) for (r,g,b,a) in pixels])
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def make_qr_svg_bytes(data: str) -> bytes:
    factory = qrcode.image.svg.SvgPathImage
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, image_factory=factory)
    qr.add_data((data or "EMPTY")); qr.make(fit=True)
    buf = io.BytesIO()
    qr.make_image().save(buf)
    return buf.getvalue()

# -------------------------------------------------
# Email Signatures (PDF with offsets)
# -------------------------------------------------

OFFSET_EN = 15
OFFSET_AR = 30

# English signature positions
EN_NAME_SIZE = 175
EN_ROLE_SIZE = 88
EN_TEXT_SIZE = 88
EN_NAME_X, EN_NAME_Y = 2900.2019, 646.9816
EN_ROLE_X, EN_ROLE_Y = 2900.2019, 856.5637
EN_EMAIL_X, EN_EMAIL_Y = 3052.323, 1245.3644
EN_WEB_X, EN_WEB_Y   = 3052.323, 1393.1144
EN_MOB_X, EN_MOB_Y   = 3052.323, 1540.8684

# Arabic signature positions
AR_NAME_SIZE = 175
AR_ROLE_SIZE = 88
AR_TEXT_SIZE = 88
AR_NAME_X, AR_NAME_Y = 2665.7742, 616.9854
AR_ROLE_X, AR_ROLE_Y = 2665.7742, 846.5675
AR_EMAIL_X, AR_EMAIL_Y = 2519.613, 1215.3663
AR_WEB_X, AR_WEB_Y   = 2519.613, 1363.1202
AR_MOB_X, AR_MOB_Y   = 2519.613, 1510.8741

def signature_en_pdf(person: dict) -> bytes:
    bg_img = Image.open(SIG_BG_EN).convert("RGBA")
    W, H = bg_img.size
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(W, H))
    c.drawImage(ImageReader(bg_img), 0, 0, W, H)
    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    c.setFont("PingBold", EN_NAME_SIZE)
    name = f"{person.get('First_Name','')} {person.get('Last_Name','')}"
    c.drawString(EN_NAME_X, H - EN_NAME_Y - OFFSET_EN, name)

    c.setFont("PingRegular", EN_ROLE_SIZE)
    c.drawString(EN_ROLE_X, H - EN_ROLE_Y - OFFSET_EN, person.get("Role", "") or "")

    c.setFont("PingRegular", EN_TEXT_SIZE)
    if person.get("Email"):  c.drawString(EN_EMAIL_X, H - EN_EMAIL_Y - OFFSET_EN, person["Email"])
    if person.get("Website"):c.drawString(EN_WEB_X,   H - EN_WEB_Y   - OFFSET_EN, person["Website"])
    if person.get("Mobile"): c.drawString(EN_MOB_X,   H - EN_MOB_Y   - OFFSET_EN, person["Mobile"])

    c.save()
    return buf.getvalue()

def signature_ar_pdf(person: dict) -> bytes:
    bg_img = Image.open(SIG_BG_AR).convert("RGBA")
    W, H = bg_img.size
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(W, H))
    c.drawImage(ImageReader(bg_img), 0, 0, W, H)
    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    name_ar_raw = person.get("Arabic_Name", "") or ""
    role_en     = person.get("Role", "") or ""
    role_ar_raw = ROLES_MAP.get(role_en, role_en)

    if name_ar_raw:
        c.setFont("PingBold", AR_NAME_SIZE)
        c.drawRightString(AR_NAME_X, H - AR_NAME_Y - OFFSET_AR, arabic_ready(name_ar_raw))
    if role_ar_raw:
        c.setFont("PingRegular", AR_ROLE_SIZE)
        c.drawRightString(AR_ROLE_X, H - AR_ROLE_Y - OFFSET_AR, arabic_ready(role_ar_raw))

    c.setFont("PingRegular", AR_TEXT_SIZE)
    if person.get("Email"):  c.drawRightString(AR_EMAIL_X, H - AR_EMAIL_Y - OFFSET_AR, person["Email"])
    if person.get("Website"):c.drawRightString(AR_WEB_X,   H - AR_WEB_Y   - OFFSET_AR, person["Website"])
    if person.get("Mobile"): c.drawRightString(AR_MOB_X,   H - AR_MOB_Y   - OFFSET_AR, person["Mobile"])

    c.save()
    return buf.getvalue()

# -------------------------------------------------
# BUSINESS CARD
# -------------------------------------------------

def business_card_pdf(person: dict) -> bytes:
    buf = io.BytesIO()
    W, H = 90*mm, 50*mm
    margin = 8*mm
    c = canvas.Canvas(buf, pagesize=(W, H))

    # Front
    c.drawImage(CARD_FRONT, 0, 0, W, H)
    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    # English
    c.setFont("PingBold", 7)
    en_name = f"{person.get('First_Name','')} {person.get('Last_Name','')}".strip()
    c.drawString(margin, H - margin - 5, en_name)

    c.setFont("PingRegular", 7)
    en_role = person.get("Role","") or ""
    c.drawString(margin, H - margin - 16.5, en_role)

    c.drawString(margin, margin + 13, person.get("Email","") or "")
    c.drawString(margin, margin,       person.get("Mobile","") or "")

    # Arabic (right aligned)
    ar_name = arabic_ready(person.get("Arabic_Name","") or "")
    ar_role = arabic_ready(ROLES_MAP.get(en_role, en_role))

    def draw_right(font_name, size, x_right, y_top_from_edge):
        c.setFont(font_name, size)
        return lambda text: c.drawString(x_right - pdfmetrics.stringWidth(text, font_name, size),
                                         H - y_top_from_edge, text)

    if ar_name:
        draw_right("PingBold", 7, W - margin, (margin + 5))(ar_name)
    if ar_role:
        draw_right("PingRegular", 7, W - margin, (margin + 16.5))(ar_role)

    # QR (transparent PNG)
    vcf = vcard_from_person(person)
    qr_png = make_qr_png_bytes(vcf, fill_color=COLOR_HEX)
    qr_img = Image.open(io.BytesIO(qr_png)).convert("RGBA")
    qr_w = 19*mm
    c.drawImage(ImageReader(qr_img), W - margin - qr_w + 1*mm, margin - 1*mm, qr_w, qr_w, mask="auto")

    # Back
    c.showPage()
    c.drawImage(CARD_BACK, 0, 0, W, H)
    c.save()
    return buf.getvalue()

# -------------------------------------------------
# ZIP WRITERS
# -------------------------------------------------

def _join_root(root: str, path: str) -> str:
    root = (root or "").strip().strip("/\\")
    return f"{root}/{path}" if root else path

def write_full_package_to_zip(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    first = p.get("First_Name","") or ""
    last  = p.get("Last_Name","")  or ""
    role  = p.get("Role","")       or ""
    comp  = p.get("Company","")    or ""
    folder = f"{first}_{last}_{role}_{comp}".replace(" ","_")

    zipf.writestr(_join_root(root, f"{folder}/Signature_EN.pdf"), signature_en_pdf(p))
    zipf.writestr(_join_root(root, f"{folder}/Signature_AR.pdf"), signature_ar_pdf(p))
    zipf.writestr(_join_root(root, f"{folder}/BusinessCard.pdf"), business_card_pdf(p))

    vcf = vcard_from_person(p)
    base = f"{first}_{last}".replace(" ","_")
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.png"), make_qr_png_bytes(vcf, fill_color=COLOR_HEX))
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.svg"), make_qr_svg_bytes(vcf))
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.vcf"), vcf)

def write_card_flat(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    base = f"{p.get('First_Name','')}_{p.get('Last_Name','')}".replace(" ","_")
    zipf.writestr(_join_root(root, f"BusinessCards_Flat/{base}.pdf"), business_card_pdf(p))

def write_signature_flat(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    base = f"{p.get('First_Name','')}_{p.get('Last_Name','')}".replace(" ","_")
    zipf.writestr(_join_root(root, f"Signatures_Flat/{base}_EN.pdf"), signature_en_pdf(p))
    zipf.writestr(_join_root(root, f"Signatures_Flat/{base}_AR.pdf"), signature_ar_pdf(p))

# -------------------------------------------------
# EXCEL TEMPLATE
# -------------------------------------------------

TEMPLATE_HEADERS = [
    "First_Name","Last_Name","Arabic_Name","Department","Role",
    "Company","Mobile","Email","Website","Location","Google_Maps_Link","Notes"
]

SAMPLE_ROWS = [
    ["Abdurrahman","Mohammed","ÿπÿ®ÿØÿßŸÑÿ±ÿ≠ŸÖŸÜ ŸÖÿ≠ŸÖÿØ","PMO","Head of Strategy","Alraedah Finance","500725242","Abdurrahman@alraedah.sa","https://www.alraedah.sa","Riyadh, Saudi","", ""],
    ["Nouf","Mohammed","ŸÜŸàŸÅ ŸÖÿ≠ŸÖÿØ","IT","Senior Analyst","Alraedah Digital","0509876543","nouf@alraedah.sa","https://www.alraedah.sa","Jeddah, Saudi","", ""],
    ["Hayat","Aldosari","ÿ≠Ÿäÿßÿ© ÿßŸÑÿØŸàÿ≥ÿ±Ÿä","Risk","Credit Underwriting Analyst","Alraedah Finance","0562223344","hayat@alraedah.sa","https://www.alraedah.sa","Dammam, Saudi","", ""],
]

def build_excel_template_bytes(include_samples: bool=True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    for c, head in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=c, value=head)
    if include_samples:
        r = 2
        for row in SAMPLE_ROWS:
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
            r += 1
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# -------------------------------------------------
# STREAMLIT APP
# -------------------------------------------------

def ensure_session_defaults():
    if "selected_department" not in st.session_state:
        st.session_state.selected_department = list(DEPARTMENTS.keys())[0]
    if "role_select" not in st.session_state:
        st.session_state.role_select = DEPARTMENTS[st.session_state.selected_department][0]

# -----------------------------
# HEADER
# -----------------------------
st.markdown(
    """
    <div style='text-align: left;'>
        <h1 style='font-weight: 700; margin-bottom: 0;'>Internal Hub</h1>
        <p style='font-size: 0.9rem; color: #5f6368; margin-top: 0.3rem;'>
            Generate official bilingual email signatures and business cards ‚Äî 
            perfectly aligned with Alraedah‚Äôs visual identity and ready for download as high-quality PDFs.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

st.divider()

# -----------------------------
# MODE FIRST (so we can branch UI)
# -----------------------------
st.subheader("Generation Mode")
st.caption("Generate for a single employee or upload an Excel for many.")
mode = st.radio(
    "Select a generation method:",
    ["Single Employee Entry", "Batch Upload (via Excel Template)"],
    index=0,
    key="mode_radio",
    horizontal=True
)

st.divider()

# -----------------------------
# OUTPUT TYPE (common to both)
# -----------------------------
st.subheader("Output Type")
st.caption("Select which assets you want to include in your ZIP file.")
download_options = st.multiselect(
    "Choose what to include:",
    ["Full Package (All Files)", "Business Cards Only", "Email Signatures Only"],
    default=["Full Package (All Files)"],
    key=f"dl_scenarios_{'single' if mode=='Single Employee Entry' else 'batch'}"
)

# -----------------------------
# BATCH-ONLY: Excel Template download
# -----------------------------
if mode == "Batch Upload (via Excel Template)":
    st.divider()
    st.subheader("Excel Template")
    st.caption(
        "Download the official Excel template to add employee details in bulk. "
        "Once filled, upload it below to instantly generate all email signatures and business cards."
    )
    include_samples = st.checkbox(
        "Include sample rows",
        value=True,
        key="tpl_samples_batch"
    )
    st.download_button(
        "Download Excel Template",
        build_excel_template_bytes(include_samples=include_samples),
        file_name="alraedah_template.xlsx",
        key="btn_download_template_batch",
        use_container_width=True
    )

# -------------------------------------------------
# SINGLE MODE
# -------------------------------------------------
if mode == "Single Employee Entry":
    ensure_session_defaults()

    st.divider()
    st.markdown("#### Employee Details")
    first_in = st.text_input("First Name", key="first_name")
    last_in  = st.text_input("Last Name",  key="last_name")
    first = normalize_name(first_in)
    last  = normalize_name(last_in)
    arabic_name = st.text_input("Arabic Name (AR)", key="arabic_name")

    st.divider()
    st.markdown("#### Department & Role")
    dept_keys = list(DEPARTMENTS.keys())
    default_dept = st.session_state.get("selected_department", dept_keys[0])

    dept = st.selectbox(
        "Department",
        dept_keys,
        index=dept_keys.index(default_dept),
        key="dep_select",
        help="Choose a department to update the role list automatically."
    )

    if dept != st.session_state.get("selected_department"):
        st.session_state.selected_department = dept
        st.session_state.role_select = DEPARTMENTS[dept][0]

    roles_for_dept = DEPARTMENTS.get(st.session_state.selected_department, [])
    if roles_for_dept:
        if st.session_state.get("role_select") not in roles_for_dept:
            st.session_state.role_select = roles_for_dept[0]
        role_index = roles_for_dept.index(st.session_state.role_select)
    else:
        st.session_state.role_select = ""
        role_index = 0

    role = st.selectbox(
        "Role",
        roles_for_dept,
        index=role_index if roles_for_dept else 0,
        key="role_select"
    )

    st.divider()
    st.markdown("#### Contact Information")
    company    = st.text_input("Company", key="company")
    mobile_raw = st.text_input("Mobile (05..., 5...., 966..., +966...)", key="mobile_raw")
    mobile_norm, _ = normalize_saudi_mobile(mobile_raw)
    email_raw  = st.text_input("Email", key="email_raw")
    email_norm, _  = normalize_email(email_raw)
    website    = st.text_input("Website", key="website")
    location   = st.text_input("Location", key="location")
    gmap_link  = st.text_input("Google Maps Link (Optional)", key="gmap")
    notes      = st.text_area("Notes", height=60, key="notes")

    submitted_single = st.button("Generate", use_container_width=True, key="btn_generate_single")
    if submitted_single:
        # Build person and ZIP
        person = {
            "First_Name": first,
            "Last_Name": last,
            "Arabic_Name": arabic_name,
            "Department": dept,
            "Role": st.session_state.role_select or "",
            "Company": company,
            "Mobile": mobile_norm,
            "Email": email_norm,
            "Website": website,
            "Location": location,
            "Google_Maps_Link": gmap_link,
            "Notes": notes
        }

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w") as zipf:
            if "Full Package (All Files)" in download_options:
                write_full_package_to_zip(zipf, person)
            if "Business Cards Only" in download_options:
                write_card_flat(zipf, person)
            if "Email Signatures Only" in download_options:
                write_signature_flat(zipf, person)
        zip_buf.seek(0)

        st.success("ZIP package ready! Click below to download.")
        st.download_button(
            "‚¨áÔ∏è Download ZIP",
            zip_buf,
            file_name="Outputs.zip",
            key="single_download_zip",
            use_container_width=True
        )

# -------------------------------------------------
# BATCH MODE
# -------------------------------------------------
elif mode == "Batch Upload (via Excel Template)":
    st.divider()
    st.markdown("#### Upload Excel Template")
    batch_file = st.file_uploader("Upload your Excel file", type=["xlsx"], key="batch_upload")

    if batch_file is not None:
        try:
            df = pd.read_excel(batch_file).fillna("")
            st.success(f"Loaded {len(df)} rows from Excel file.")
        except Exception as e:
            st.error(f"Failed to read Excel file: {e}")
            st.stop()

        # Data cleaning / mapping
        mapped = df.rename(columns={
            "First Name": "First_Name",
            "Last Name": "Last_Name",
            "Arabic Name (AR)": "Arabic_Name",
            "Department": "Department",
            "Role": "Role",
            "Company": "Company",
            "Mobile": "Mobile",
            "Email": "Email",
            "Website": "Website",
            "Location": "Location",
            "Google Maps Link (Optional)": "Google_Maps_Link",
            "Notes": "Notes"
        })

        # Normalize
        mapped["Mobile_Normalized"], mapped["Mobile_Valid"] = zip(*mapped["Mobile"].map(normalize_saudi_mobile))
        mapped["Email_Normalized"],  mapped["Email_Valid"]  = zip(*mapped["Email"].map(normalize_email))

        # Validation & dedupe
        total = len(mapped)
        dup_mobile = mapped["Mobile_Normalized"].duplicated(keep=False)
        dup_email  = mapped["Email_Normalized"].duplicated(keep=False)
        duplicates = mapped[dup_mobile | dup_email]
        invalids   = mapped[~mapped["Email_Valid"] | ~mapped["Mobile_Valid"]]
        clean_df   = mapped[~mapped.index.isin(duplicates.index) & ~mapped.index.isin(invalids.index)]

        clean_count     = len(clean_df)
        invalid_count   = len(invalids)
        duplicate_count = len(duplicates)

        # Summary card
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8fafc, #eef3fa);
                    border: 1px solid #e0e6f0; border-radius: 12px; padding: 18px 20px;
                    margin: 15px 0; box-shadow: 0 2px 6px rgba(0,0,0,0.03);">
          <div style="font-size:17px; font-weight:600; color:#254489; margin-bottom:6px;">üìä Batch Summary</div>
          <div style="font-size:15px; color:#333;">
            <span style="font-weight:500;">Total:</span> {total} employees<br>
            <span style="color:#254489; font-weight:500;">‚úÖ Clean:</span> {clean_count}<br>
            <span style="color:#9e9e9e; font-weight:500;">‚ö†Ô∏è Duplicates:</span> {duplicate_count}<br>
            <span style="color:#cc3333; font-weight:500;">‚ùå Invalid:</span> {invalid_count}
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.dataframe(
            mapped[["First_Name","Last_Name","Arabic_Name","Department","Role","Company"]],
            use_container_width=True
        )

        # Custom ZIP name (prefix only; date appended automatically)
        today_str = datetime.now().strftime("%Y%m%d")
        custom_zip_prefix = st.text_input(
            "üì¶ Enter ZIP Name Prefix",
            value="Batch_Outputs",
            key="batch_zip_prefix",
            help="Type the prefix for your ZIP file (the date will be added automatically)."
        )
        safe_root   = re.sub(r"[^A-Za-z0-9._-]+", "_", custom_zip_prefix).strip("_") or "Batch_Outputs"
        zip_filename = f"{safe_root}_{today_str}.zip"

        # Generate ZIP
        if st.button("Generate ZIP Outputs", key="btn_batch_generate", use_container_width=True):
            if clean_count == 0:
                st.warning("No valid employees to generate ZIP files. Please fix invalid entries and try again.")
                st.stop()

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zipf:
                # Summary file
                summary_lines = [
                    "üìä Batch Summary Report",
                    "=======================",
                    f"Total Employees: {total}",
                    f"Clean: {clean_count} ‚úÖ",
                    f"Invalid: {invalid_count} ‚ùå",
                    f"Duplicates: {duplicate_count} ‚ö†Ô∏è",
                    "",
                    f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    ""
                ]
                zipf.writestr(f"{safe_root}/Batch_Summary.txt", "\n".join(summary_lines))

                # Each employee‚Äôs files
                for _, row in clean_df.iterrows():
                    person = {
                        "First_Name": row.get("First_Name", ""),
                        "Last_Name": row.get("Last_Name", ""),
                        "Arabic_Name": row.get("Arabic_Name", ""),
                        "Department": row.get("Department", ""),
                        "Role": row.get("Role", ""),
                        "Company": row.get("Company", ""),
                        "Mobile": row.get("Mobile_Normalized", ""),
                        "Email": row.get("Email_Normalized", ""),
                        "Website": row.get("Website", ""),
                        "Location": row.get("Location", ""),
                        "Notes": row.get("Notes", "")
                    }

                    folder_name = f"{safe_root}/{person['First_Name']}_{person['Last_Name']}"
                    if "Full Package (All Files)" in download_options:
                        write_full_package_to_zip(zipf, person, root=folder_name)
                    if "Business Cards Only" in download_options:
                        write_card_flat(zipf, person, root=folder_name)
                    if "Email Signatures Only" in download_options:
                        write_signature_flat(zipf, person, root=folder_name)

            zip_buf.seek(0)
            st.success("ZIP generation completed successfully!")
            st.download_button(
                f"‚¨áÔ∏è Download {zip_filename}",
                zip_buf,
                file_name=zip_filename,
                key="batch_download_zip",
                use_container_width=True
            )

# -------------------------------------------------
# FOOTER
# -------------------------------------------------
st.markdown(
    """
    <hr style="margin-top:3em; margin-bottom:0.5em; border: 0; border-top: 1px solid #e0e0e0;">
    <div style="text-align:center; color:gray; font-size:13px;">
        ¬© Alraedah Finance ‚Äî Internal Use Only
    </div>
    """,
    unsafe_allow_html=True
)

# -------------------------------------------------
# SIDEBAR
# -------------------------------------------------
with st.sidebar:
    st.markdown("### About Internal Hub")
    st.caption(
        "An internal tool developed for **Alraedah Finance** to automatically generate "
        "bilingual (English & Arabic) email signatures and business cards that align "
        "with the company‚Äôs brand identity."
    )
    st.divider()
    st.markdown("### How It Works")
    st.markdown(
        """
        1. **Single Entry** ‚Üí Generate materials for one employee.  
        2. **Batch Upload** ‚Üí Upload an Excel file with multiple employees.  
        3. **Download ZIP** ‚Üí Get all signatures, cards, and QR files instantly.
        """
    )
    st.divider()
    st.markdown("### Output Files")
    st.markdown(
        """
        - Email Signature (EN / AR)  
        - Business Card (Front / Back)  
        - QR Code (PNG / SVG / VCF)  
        """
    )
    st.divider()
    st.caption("Version 1.0.1 ¬∑ Built by Abdurrahman Alowain ¬∑ For Internal Use Only")