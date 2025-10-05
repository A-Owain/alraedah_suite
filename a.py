import io
import os
import re
import zipfile
from datetime import datetime

import streamlit as st
import pandas as pd

from PIL import Image, ImageDraw, ImageFont
import qrcode, qrcode.image.svg

# ReportLab (PDF)
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Arabic shaping/bidi
import arabic_reshaper
from bidi.algorithm import get_display

# Excel template
from openpyxl import Workbook

# -------------------------------------------------
# GLOBAL CONFIG
# -------------------------------------------------

APP_TITLE = "Internal Hub"

# Fonts (must be valid TTFs with Arabic glyphs)
FONT_BOLD_PATH = "assets/Fonts/PingAR+LT-Bold.ttf"
FONT_REGULAR_PATH = "assets/Fonts/PingAR+LT-Regular.ttf"

# Backgrounds for signatures
SIG_BG_EN = "assets/images/eng-bg.png"
SIG_BG_AR = "assets/images/arbc-bg.png"

# Business card faces
CARD_FRONT = "assets/Images/front.png"
CARD_BACK  = "assets/Images/back.png"

# Brand color
COLOR_HEX = "#254489"
COLOR_RGB_REPORTLAB = (0.145, 0.266, 0.537)  # #254489

# Register fonts for ReportLab (used by PDFs)
pdfmetrics.registerFont(TTFont("PingBold",    FONT_BOLD_PATH))
pdfmetrics.registerFont(TTFont("PingRegular", FONT_REGULAR_PATH))

# -------------------------------------------------
# DEPARTMENTS & ROLES
# -------------------------------------------------

DEPARTMENTS = {
    "PMO": [
        "Head of Strategy","Head of PMO","IT Program Manager","Business Analyst", "Board Member & CEO"
    ],
    "IT": [
        "Head of IT","Senior Manager of Technology Operation & Applications",
        "Senior Technical & Network Engineer","Senior DevOps Engineer",
        "Senior Analyst","IT Support Administrator","Outsourced"
    ],
    "Engineering & Data": [
        "Head of Engineering","Head of Data","Senior Software Engineer",
        "Staff IOS Engineer","Data Engineer","Associate Data Engineer",
        "Data Analyst","Senior Analyst","Analyst","Outsourced"
    ],
    "Product": ["Principal","Lead","Senior"],
    "Cybersecurity": ["CISO","Cybersecurity Engineer","Cybersecurity Analyst","Threat Intelligence Analyst","Outsourced"],
    "Risk": [
        "Head of Risk","AVP - Operations Manager","AVP","Credit Manager",
        "Senior Associate","Senior Operations Analyst","Senior Analyst",
        "Credit Underwriting Senior Analyst","Credit Underwriting Analyst",
        "Credit Administration Analyst","Operations Analyst","Associate","Analyst"
    ],
    "Legal": [
        "Senior Corporate Government Analyst","Collection Manager","Senior Legal Analyst",
        "Senior Associate","Team Leader","Legal Analyst","Senior Analyst",
        "Collection Officer","Associate","Analyst","Outsourced"
    ],
    "Customer Care": ["Head of Customer Care","Senior Customer Care Specialist","Customer Care Specialist"],
    "Compliance": ["VP - Head of Compliance","Senior AML Officer","AML & Compliance Officer","Compliance Specialist"],
    "Digital Lending": ["Commercial Director","Digital Marketing Manager","Customer Success Manager","Customer Success Specialist"],
    "Direct Lending": [
        "Head of Sales","Senior Key Accounts Manager","Key Accounts Manager","Sales Manager",
        "Senior Team Leader","Team Leader","Growth Manager","Growth Consultant II",
        "Growth Consultant I","Growth Associate","Sales Operations","Business Development Representative",
        "Quality Assurance Specialist"
    ],
    "Marketing": ["Senior Communication Specialist","Marketing Specialist","Social Media Specialist","Marketing Executive"],
    "Finance": ["VP - Head of Finance","VP - Finance","Senior Finance Analyst","Senior Financial Associate","Financial Analyst","Financial Associate"]
}

ROLES_MAP = {
    "Head of Strategy":"رئيس الاستراتيجية","Head of PMO":"رئيس مكتب إدارة المشاريع","IT Program Manager":"مدير برنامج تقنية المعلومات",
    "Business Analyst":"محلل أعمال","Head of IT":"رئيس تقنية المعلومات","Board Member & CEO":"عضو مجلس الإدارة والرئيس التنفيذي",
    "Senior Manager of Technology Operation & Applications":"مدير أول لعمليات التقنية والتطبيقات",
    "Senior Analyst":"محلل أول","Outsourced":"موظف خارجي","Senior Technical & Network Engineer":"مهندس تقني وشبكات أول",
    "IT Support Administrator":"مسؤول دعم تقنية المعلومات","Senior DevOps Engineer":"مهندس ديف أوبس أول",
    "Head of Engineering":"رئيس الهندسة","Senior Software Engineer":"مهندس برمجيات أول","Staff IOS Engineer":"مهندس نظم iOS",
    "Head of Data":"رئيس البيانات","Data Engineer":"مهندس بيانات","Analyst":"محلل","Associate Data Engineer":"مهندس بيانات مشارك",
    "Data Analyst":"محلل بيانات","Principal":"رئيس قسم","Lead":"قائد فريق","Senior":"كبير","CISO":"رئيس أمن المعلومات",
    "Cybersecurity Engineer":"مهندس أمن سيبراني","Cybersecurity Analyst":"محلل أمن سيبراني","Threat Intelligence Analyst":"محلل استخبارات التهديدات",
    "Head of Risk":"رئيس إدارة المخاطر","Risk Officer":"مسؤول مخاطر","AVP":"مساعد نائب الرئيس","Credit Manager":"مدير الائتمان",
    "Credit Underwriting Senior Analyst":"محلل أول للاكتتاب الائتماني","Credit Underwriting Analyst":"محلل اكتتاب ائتماني",
    "AVP - Operations Manager":"مساعد نائب الرئيس - مدير العمليات","Senior Associate":"مساعد أول","Associate":"مساعد",
    "Senior Operations Analyst":"محلل عمليات أول","Operations Analyst":"محلل عمليات","Credit Administration Analyst":"محلل إدارة الائتمان",
    "Senior Legal Analyst":"محلل قانوني أول","Legal Analyst":"محلل قانوني","Team Leader":"قائد فريق","Collection Manager":"مدير التحصيل",
    "Senior Collection Officer":"مسؤول تحصيل أول","Collection Officer":"مسؤول تحصيل","Senior Corporate Government Analyst":"محلل حوكمة مؤسسية أول",
    "Head of Customer Care":"رئيس خدمة العملاء","Senior Customer Care Specialist":"أخصائي خدمة العملاء أول","Customer Care Specialist":"أخصائي خدمة العملاء",
    "VP - Head of Compliance":"نائب الرئيس - رئيس الامتثال","Senior AML Officer":"مسؤول أول لمكافحة غسل الأموال",
    "AML & Compliance Officer":"مسؤول امتثال ومكافحة غسل الأموال","Compliance Specialist":"أخصائي امتثال",
    "Commercial Director":"المدير التجاري","Digital Marketing Manager":"مدير التسويق الرقمي","Customer Success Manager":"مدير نجاح العملاء",
    "Customer Success Specialist":"أخصائي نجاح العملاء","Head of Sales":"رئيس المبيعات","Sales Manager":"مدير مبيعات","Senior Team Leader":"قائد فريق أول",
    "Growth Manager":"مدير النمو","Growth Consultant II":"مستشار نمو II","Growth Consultant I":"مستشار نمو I","Growth Associate":"مساعد نمو",
    "Key Accounts Manager":"مدير الحسابات الرئيسية","Senior Key Accounts Manager":"مدير الحسابات الرئيسية أول","Sales Operations":"عمليات المبيعات",
    "Business Development Representative":"ممثل تطوير الأعمال","Quality Assurance Specialist":"أخصائي ضمان الجودة",
    "Senior Communication Specialist":"أخصائي اتصالات أول","Marketing Executive":"تنفيذي تسويق","Social Media Specialist":"أخصائي وسائل التواصل الاجتماعي",
    "Marketing Specialist":"أخصائي تسويق","VP - Head of Finance":"نائب الرئيس - رئيس المالية","VP - Finance":"نائب الرئيس - المالية",
    "Senior Financial Associate":"مساعد مالي أول","Financial Associate":"مساعد مالي","Senior Finance Analyst":"محلل مالي أول","Financial Analyst":"محلل مالي"
}

# -------------------------------------------------
# HELPERS
# -------------------------------------------------

def arabic_ready(text: str) -> str:
    """Shape + bidi Arabic so it displays connected (not mirrored)."""
    return get_display(arabic_reshaper.reshape(text)) if text else ""

def normalize_saudi_mobile(mobile: str) -> tuple[str, bool]:
    if not mobile: return ("", False)
    clean = re.sub(r"\D", "", str(mobile))
    if clean.startswith("00966") and len(clean) == 14: clean = clean[2:]
    if clean.startswith("9665") and len(clean) == 12: return (f"+{clean}", True)
    if clean.startswith("05")   and len(clean) == 10: return (f"+966{clean[1:]}", True)
    if clean.startswith("5")    and len(clean) == 9:  return (f"+966{clean}", True)
    if clean.startswith("966")  and len(clean) == 12: return (f"+{clean}", True)
    if clean.startswith("+966") and len(clean) == 13: return (clean, True)
    return (mobile, False)

def normalize_email(email: str) -> tuple[str, bool]:
    if not email: return ("", False)
    e = email.strip().lower()
    return (e, "@" in e and e.count("@") == 1)

def normalize_name(name: str) -> str:
    return name.strip().capitalize() if name else ""

def vcard_from_person(p: dict) -> str:
    first = p.get("First_Name","") or ""
    last  = p.get("Last_Name","")  or ""
    org   = p.get("Company","")    or ""
    title = p.get("Role","")       or ""
    tel   = p.get("Mobile","")     or ""
    email = p.get("Email","")      or ""
    url   = p.get("Website","")    or ""
    loc   = p.get("Location","")   or ""
    notes = p.get("Notes","")      or ""
    return f"""BEGIN:VCARD
VERSION:3.0
N:{last};{first}
FN:{first} {last}
ORG:{org}
TITLE:{title}
TEL;TYPE=CELL:{tel}
EMAIL:{email}
URL:{url}
ADR;TYPE=WORK:{loc}
NOTE:{notes}
END:VCARD"""

def make_qr_png_bytes(data: str, fill_color="#254489") -> bytes:
    """Transparent-background QR PNG."""
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(data or "")
    qr.make(fit=True)
    img = qr.make_image(fill_color=fill_color, back_color="white").convert("RGBA")
    # strip white to transparency
    pixels = img.getdata()
    img.putdata([(r,g,b,0) if r>240 and g>240 and b>240 else (r,g,b,a) for (r,g,b,a) in pixels])
    buf = io.BytesIO(); img.save(buf, format="PNG"); return buf.getvalue()

def make_qr_svg_bytes(data: str) -> bytes:
    factory = qrcode.image.svg.SvgPathImage
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, image_factory=factory)
    qr.add_data((data or "EMPTY")); qr.make(fit=True)
    buf = io.BytesIO(); qr.make_image().save(buf); return buf.getvalue()

# -------------------------------------------------
# Email Signatures (PDF with offsets)
# -------------------------------------------------

OFFSET_EN = 15   # adjust until EN lines match PNG
OFFSET_AR = 30   # adjust until AR lines match PNG

# -------------------------------------------------
# SIGNATURE DIMENSIONS (exact coordinates + sizes)
# -------------------------------------------------

# English signature
EN_NAME_SIZE = 175
EN_ROLE_SIZE = 88
EN_TEXT_SIZE = 88
EN_NAME_X, EN_NAME_Y = 2900.2019, 646.9816
EN_ROLE_X, EN_ROLE_Y = 2900.2019, 856.5637
EN_EMAIL_X, EN_EMAIL_Y = 3052.323, 1245.3644
EN_WEB_X, EN_WEB_Y = 3052.323, 1393.1144
EN_MOB_X, EN_MOB_Y = 3052.323, 1540.8684

# Arabic signature
AR_NAME_SIZE = 175
AR_ROLE_SIZE = 88
AR_TEXT_SIZE = 88
AR_NAME_X, AR_NAME_Y = 2665.7742, 616.9854
AR_ROLE_X, AR_ROLE_Y = 2665.7742, 846.5675
AR_EMAIL_X, AR_EMAIL_Y = 2519.613, 1215.3663
AR_WEB_X, AR_WEB_Y = 2519.613, 1363.1202
AR_MOB_X, AR_MOB_Y = 2519.613, 1510.8741

def signature_en_pdf(person: dict) -> bytes:
    """English signature with image background + offset fix."""
    bg_img = Image.open(SIG_BG_EN).convert("RGBA")
    W, H = bg_img.size
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(W, H))
    c.drawImage(ImageReader(bg_img), 0, 0, W, H)

    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    # Name
    c.setFont("PingBold", EN_NAME_SIZE)
    name = f"{person.get('First_Name','')} {person.get('Last_Name','')}"
    c.drawString(EN_NAME_X, H - EN_NAME_Y - OFFSET_EN, name)

    # Role
    c.setFont("PingRegular", EN_ROLE_SIZE)
    c.drawString(EN_ROLE_X, H - EN_ROLE_Y - OFFSET_EN, person.get("Role", "") or "")

    # Contacts
    c.setFont("PingRegular", EN_TEXT_SIZE)
    if person.get("Email"):
        c.drawString(EN_EMAIL_X, H - EN_EMAIL_Y - OFFSET_EN, person["Email"])
    if person.get("Website"):
        c.drawString(EN_WEB_X, H - EN_WEB_Y - OFFSET_EN, person["Website"])
    if person.get("Mobile"):
        c.drawString(EN_MOB_X, H - EN_MOB_Y - OFFSET_EN, person["Mobile"])

    c.save()
    return buf.getvalue()

def signature_ar_pdf(person: dict) -> bytes:
    """Arabic signature with image background + offset fix."""
    bg_img = Image.open(SIG_BG_AR).convert("RGBA")
    W, H = bg_img.size
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(W, H))
    c.drawImage(ImageReader(bg_img), 0, 0, W, H)

    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    # Arabic-prepared strings
    name_ar_raw = person.get("Arabic_Name", "") or ""
    role_en = person.get("Role", "") or ""
    role_ar_raw = ROLES_MAP.get(role_en, role_en)

    # Directly draw selectable Arabic
    if name_ar_raw:
        c.setFont("PingBold", AR_NAME_SIZE)
        c.drawRightString(AR_NAME_X, H - AR_NAME_Y - OFFSET_AR, arabic_ready(name_ar_raw))
    if role_ar_raw:
        c.setFont("PingRegular", AR_ROLE_SIZE)
        c.drawRightString(AR_ROLE_X, H - AR_ROLE_Y - OFFSET_AR, arabic_ready(role_ar_raw))

    # Contacts (EN, right-aligned)
    c.setFont("PingRegular", AR_TEXT_SIZE)
    if person.get("Email"):
        c.drawRightString(AR_EMAIL_X, H - AR_EMAIL_Y - OFFSET_AR, person["Email"])
    if person.get("Website"):
        c.drawRightString(AR_WEB_X, H - AR_WEB_Y - OFFSET_AR, person["Website"])
    if person.get("Mobile"):
        c.drawRightString(AR_MOB_X, H - AR_MOB_Y - OFFSET_AR, person["Mobile"])

    c.save()
    return buf.getvalue()

# -------------------------------------------------
# BUSINESS CARD (Arabic selectable via same approach)
# -------------------------------------------------

def business_card_pdf(person: dict) -> bytes:
    buf = io.BytesIO()
    W, H = 90*mm, 50*mm
    margin = 8*mm
    c = canvas.Canvas(buf, pagesize=(W, H))

    # Front
    c.drawImage(CARD_FRONT, 0, 0, W, H)
    c.setFillColorRGB(*COLOR_RGB_REPORTLAB)

    # English top-left
    c.setFont("PingBold", 7)
    en_name = f"{person.get('First_Name','')} {person.get('Last_Name','')}".strip()
    c.drawString(margin, H - margin - 5, en_name)

    c.setFont("PingRegular", 7)
    en_role = person.get("Role","") or ""
    c.drawString(margin, H - margin - 16.5, en_role)

    # English contacts bottom-left
    c.drawString(margin, margin + 13, person.get("Email","") or "")
    c.drawString(margin, margin,       person.get("Mobile","") or "")

    # Arabic top-right (selectable)
    ar_name = arabic_ready(person.get("Arabic_Name","") or "")
    ar_role = arabic_ready(ROLES_MAP.get(en_role, en_role))

    # Right align by measuring width
    def draw_right(font_name, size, x_right, y_top_from_edge):
        c.setFont(font_name, size)
        return lambda text: c.drawString(x_right - pdfmetrics.stringWidth(text, font_name, size),
                                         H - y_top_from_edge, text)

    if ar_name:
        draw_right("PingBold", 7, W - margin, (margin + 5))(ar_name)
    if ar_role:
        draw_right("PingRegular", 7, W - margin, (margin + 16.5))(ar_role)

    # QR bottom-right (transparent background PNG)
    vcf = vcard_from_person(person)
    qr_png = make_qr_png_bytes(vcf, fill_color=COLOR_HEX)
    qr_img = Image.open(io.BytesIO(qr_png)).convert("RGBA")
    qr_w = 19*mm
    c.drawImage(ImageReader(qr_img), W - margin - qr_w + 1*mm, margin - 1*mm, qr_w, qr_w, mask="auto")

    # Back
    c.showPage()
    c.drawImage(CARD_BACK, 0, 0, W, H)
    c.save()
    return buf.getvalue()

# -------------------------------------------------
# ZIP WRITERS (with optional root prefix for batch)
# -------------------------------------------------

def _join_root(root: str, path: str) -> str:
    root = (root or "").strip().strip("/\\")
    return f"{root}/{path}" if root else path

def write_full_package_to_zip(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    first = p.get("First_Name","") or ""
    last  = p.get("Last_Name","")  or ""
    role  = p.get("Role","")       or ""
    comp  = p.get("Company","")    or ""
    folder = f"{first}_{last}_{role}_{comp}".replace(" ","_")

    zipf.writestr(_join_root(root, f"{folder}/Signature_EN.pdf"), signature_en_pdf(p))
    zipf.writestr(_join_root(root, f"{folder}/Signature_AR.pdf"), signature_ar_pdf(p))
    zipf.writestr(_join_root(root, f"{folder}/BusinessCard.pdf"), business_card_pdf(p))

    vcf = vcard_from_person(p)
    base = f"{first}_{last}".replace(" ","_")
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.png"), make_qr_png_bytes(vcf, fill_color=COLOR_HEX))
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.svg"), make_qr_svg_bytes(vcf))
    zipf.writestr(_join_root(root, f"{folder}/QR/{base}.vcf"), vcf)

def write_card_flat(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    base = f"{p.get('First_Name','')}_{p.get('Last_Name','')}".replace(" ","_")
    zipf.writestr(_join_root(root, f"BusinessCards_Flat/{base}.pdf"), business_card_pdf(p))

def write_signature_flat(zipf: zipfile.ZipFile, p: dict, root: str="") -> None:
    base = f"{p.get('First_Name','')}_{p.get('Last_Name','')}".replace(" ","_")
    zipf.writestr(_join_root(root, f"Signatures_Flat/{base}_EN.pdf"), signature_en_pdf(p))
    zipf.writestr(_join_root(root, f"Signatures_Flat/{base}_AR.pdf"), signature_ar_pdf(p))

# -------------------------------------------------
# EXCEL TEMPLATE
# -------------------------------------------------

TEMPLATE_HEADERS = [
    "First_Name","Last_Name","Arabic_Name","Department","Role",
    "Company","Mobile","Email","Website","Location","Google_Maps_Link","Notes"
]

SAMPLE_ROWS = [
    ["Abdurrahman","Mohammed","عبدالرحمن محمد","PMO","Head of Strategy","Alraedah Finance","500725242","Abdurrahman@alraedah.sa","https://www.alraedah.sa","Riyadh, Saudi","", ""],
    ["Nouf","Mohammed","نوف محمد","IT","Senior Analyst","Alraedah Digital","0509876543","nouf@alraedah.sa","https://www.alraedah.sa","Jeddah, Saudi","", ""],
    ["Hayat","Aldosari","حياة الدوسري","Risk","Credit Underwriting Analyst","Alraedah Finance","0562223344","hayat@alraedah.sa","https://www.alraedah.sa","Dammam, Saudi","", ""],
]

def build_excel_template_bytes(include_samples: bool=True) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Employees"
    for c, head in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=c, value=head)
    if include_samples:
        r = 2
        for row in SAMPLE_ROWS:
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
            r += 1
    out = io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()

# -------------------------------------------------
# STREAMLIT STATE HELPERS (dependent dropdown)
# -------------------------------------------------

def ensure_session_defaults():
    if "selected_department" not in st.session_state:
        st.session_state.selected_department = list(DEPARTMENTS.keys())[0]
    if "selected_role" not in st.session_state:
        st.session_state.selected_role = DEPARTMENTS[st.session_state.selected_department][0]

# -----------------------------
# HEADER SECTION (Public-ready)
# -----------------------------
st.markdown(
    """
    <div style='text-align: left;'>
        <h1 style='font-weight: 700; margin-bottom: 0;'>Internal Hub</h1>
        <p style='font-size: 0.9rem; color: #5f6368; margin-top: 0.3rem;'>
            Generate official bilingual email signatures and business cards — 
            perfectly aligned with Alraedah’s visual identity and ready for download as high-quality PDFs.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# --- Template download section ---
st.divider()
st.subheader("Excel Template")
st.caption(
    "Download the official Excel template to add employee details in bulk. "
    "Once filled, upload it below to instantly generate all email signatures and business cards — "
    "accurate, bilingual, and ready to use."
)
include_samples = st.checkbox("Include sample rows", value=True, key="tpl_samples")

st.download_button(
    "Download Excel Template",
    build_excel_template_bytes(include_samples=include_samples),
    file_name="alraedah_template.xlsx",
    key="btn_download_template",
    use_container_width=True
)

st.divider()
st.subheader("Generation Mode")
st.caption(
    "Choose how you’d like to generate materials — either for a single employee "
    "or for multiple employees using the Excel template."
)

mode = st.radio(
    "Select a generation method:",
    ["Single Employee Entry", "Batch Upload (via Excel Template)"],
    index=0,
    key="mode_radio"
)

st.divider()
st.subheader("Output Type")
st.caption("Select which assets you want to include in your ZIP file.")

download_options = st.multiselect(
    "Choose what to include:",
    ["Full Package (All Files)", "Business Cards Only", "Email Signatures Only"],
    default=["Full Package (All Files)"],
    key="dl_scenarios"
)

# ---------- SINGLE ----------
if mode == "Single Employee Entry":
    ensure_session_defaults()

    st.markdown("#### Employee Details")
    with st.form("single_form"):
        # --- Basic info ---
        first_in = st.text_input("First Name")
        last_in  = st.text_input("Last Name")
        first = normalize_name(first_in)
        last  = normalize_name(last_in)
        arabic_name = st.text_input("Arabic Name (AR)")

        st.divider()
        st.markdown("#### Department & Role")

        # Department (standalone select)
        dept = st.selectbox(
            "Department",
            list(DEPARTMENTS.keys()),
            index=list(DEPARTMENTS.keys()).index(st.session_state.selected_department),
            key="dep_select"
        )

        # ✨ Trick: refresh button inside form to update roles manually
        refresh_roles = st.form_submit_button("↻ Refresh Roles", use_container_width=False)

        # Handle role logic
        if refresh_roles:
            st.session_state.selected_department = dept
            st.session_state.selected_role = DEPARTMENTS[dept][0]

        roles_for_dept = DEPARTMENTS.get(st.session_state.selected_department, [])
        current_role = st.session_state.get("selected_role", roles_for_dept[0])
        if current_role not in roles_for_dept:
            current_role = roles_for_dept[0]
            st.session_state.selected_role = current_role

        role = st.selectbox(
            "Role",
            roles_for_dept,
            index=roles_for_dept.index(current_role),
            key="role_select"
        )
        st.session_state.selected_role = role

        st.divider()
        st.markdown("#### Contact Information")

        company = st.text_input("Company")
        mobile_raw = st.text_input("Mobile (05..., 5...., 966..., +966...)")
        mobile_norm, _ = normalize_saudi_mobile(mobile_raw)
        email_raw = st.text_input("Email")
        email_norm, _ = normalize_email(email_raw)
        website = st.text_input("Website")
        location = st.text_input("Location")
        gmap_link = st.text_input("Google Maps Link (Optional)")
        notes = st.text_area("Notes", height=60)

        submitted = st.form_submit_button("Generate", use_container_width=True)

    # --- After submission ---
    if submitted:
        person = {
            "First_Name": first, "Last_Name": last, "Arabic_Name": arabic_name,
            "Department": dept, "Role": role, "Company": company,
            "Mobile": mobile_norm, "Email": email_norm, "Website": website,
            "Location": location, "Google_Maps_Link": gmap_link, "Notes": notes
        }

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w") as zipf:
            if "Full Package (All Files)" in download_options:
                write_full_package_to_zip(zipf, person)
            if "Business Cards Only" in download_options:
                write_card_flat(zipf, person)
            if "Email Signatures Only" in download_options:
                write_signature_flat(zipf, person)
        zip_buf.seek(0)

        st.download_button(
            "⬇️ Download ZIP",
            zip_buf,
            file_name="Outputs.zip",
            key="single_download_zip",
            use_container_width=True
        )

# ---------- BATCH ----------
else:
    uploaded = st.file_uploader("Upload Excel (.xlsx) with Employees sheet", type=["xlsx"], key="uploader_xlsx")

    # NEW: custom batch root folder name
    default_batch_name = f"Batch_Outputs_{datetime.now().strftime('%Y%m%d')}"
    batch_root_name = st.text_input("Batch Folder Name (Optional)", value=default_batch_name, help="Used as the ZIP filename and the root folder inside the ZIP.")

    if uploaded:
        try: df = pd.read_excel(uploaded, sheet_name="Employees")
        except Exception: df = pd.read_excel(uploaded)

        aliases = {
            "First_Name": ["First_Name","First Name","First","FName"],
            "Last_Name":  ["Last_Name","Last Name","Last","LName"],
            "Arabic_Name":["Arabic_Name","Arabic Name","Name_AR","AR_Name"],
            "Department": ["Department","Dept"],
            "Role":       ["Role","Position","Title"],
            "Company":    ["Company","Org","Organization"],
            "Mobile":     ["Mobile","Phone","Cell","Mobile_Normalized"],
            "Email":      ["Email","E-mail","Mail","Email_Normalized"],
            "Website":    ["Website","Site","URL"],
            "Location":   ["Location","Address","City"],
            "Google_Maps_Link":["Google_Maps_Link","Maps","Google Maps Link","GMaps"],
            "Notes":      ["Notes","Remark","Comment"],
        }
        def resolve_col(frame, cands, default=None):
            for c in cands:
                if c in frame.columns: return c
            return default

        mapped = pd.DataFrame()
        for k, cands in aliases.items():
            col = resolve_col(df, cands)
            mapped[k] = df[col] if col else ""

        mapped["First_Name"] = mapped["First_Name"].apply(normalize_name)
        mapped["Last_Name"]  = mapped["Last_Name"].apply(normalize_name)

        # Normalize contacts
        em_n, em_ok = [], []
        for v in mapped["Email"].tolist():
            n, ok = normalize_email(v); em_n.append(n); em_ok.append(ok)
        mapped["Email_Normalized"], mapped["Email_Valid"] = em_n, em_ok

        mb_n, mb_ok = [], []
        for v in mapped["Mobile"].tolist():
            n, ok = normalize_saudi_mobile(v); mb_n.append(n); mb_ok.append(ok)
        mapped["Mobile_Normalized"], mapped["Mobile_Valid"] = mb_n, mb_ok

        dup_email  = mapped["Email_Normalized"].duplicated(Keep=False) if hasattr(mapped["Email_Normalized"], 'duplicated') else mapped["Email_Normalized"].duplicated(keep=False)
        dup_email  = dup_email & mapped["Email_Normalized"].ne("")
        dup_mobile = mapped["Mobile_Normalized"].duplicated(keep=False) & mapped["Mobile_Normalized"].ne("")
        duplicates = mapped[dup_email | dup_mobile].copy()
        invalids   = mapped[(~mapped["Email_Valid"]) | (~mapped["Mobile_Valid"])].copy()

        total=len(mapped); invalid_count=len(invalids); duplicate_count=len(duplicates)
        clean_count = total - len(invalids.index.union(duplicates.index))

        st.info(f"📊 Batch Summary: {total} rows → {clean_count} clean ✅ | {invalid_count} invalid ❌ | {duplicate_count} duplicates ⚠️")
        st.dataframe(mapped[["First_Name","Last_Name","Arabic_Name","Department","Role",
                             "Company","Email_Normalized","Email_Valid",
                             "Mobile_Normalized","Mobile_Valid","Website","Location","Notes"]])

        if st.button("Generate ZIP Outputs", key="btn_batch_generate", use_container_width=True):
            safe_root = re.sub(r"[^A-Za-z0-9._-]+", "_", (batch_root_name or default_batch_name)).strip("_") or default_batch_name

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w") as zipf:
                # reports
                summary = [
                    "📊 Batch Summary Report","=======================",
                    f"Total Employees: {total}",
                    f"Clean: {clean_count} ✅", f"Invalid: {invalid_count} ❌", f"Duplicates: {duplicate_count} ⚠️",""
                ]
                if invalid_count:
                    summary.append("Invalid Entries:")
                    for _, r in invalids.iterrows():
                        who = f"{r.get('First_Name','')} {r.get('Last_Name','')}".strip()
                        reasons=[]
                        if not r.get("Email_Valid",True): reasons.append("Email")
                        if not r.get("Mobile_Valid",True): reasons.append("Mobile")
                        summary.append(f"- {who}: {', '.join(reasons) or 'Unknown'}")
                    summary.append("")
                if duplicate_count:
                    summary.append("Duplicate Entries (by normalized Email/Mobile):")
                    for _, r in duplicates.iterrows():
                        who = f"{r.get('First_Name','')} {r.get('Last_Name','')}".strip()
                        summary.append(f"- {who} | Email: {r.get('Email_Normalized','')} | Mobile: {r.get('Mobile_Normalized','')}")
                    summary.append("")
                summary.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

                # write reports under root
                zipf.writestr(_join_root(safe_root, "Batch_Summary.txt"), "\n".join(summary))
                if invalid_count:
                    s = io.StringIO(); invalids.to_csv(s, index=False)
                    zipf.writestr(_join_root(safe_root, "Reports/Invalid_Entries.csv"), s.getvalue())
                if duplicate_count:
                    s = io.StringIO(); duplicates.to_csv(s, index=False)
                    zipf.writestr(_join_root(safe_root, "Reports/Duplicates.csv"), s.getvalue())

                # outputs
                for _, row in mapped.iterrows():
                    person = {
                        "First_Name": row.get("First_Name","") or "",
                        "Last_Name":  row.get("Last_Name","")  or "",
                        "Arabic_Name":row.get("Arabic_Name","")or "",
                        "Department": row.get("Department","") or "",
                        "Role":       row.get("Role","")       or "",
                        "Company":    row.get("Company","")    or "",
                        "Mobile":     row.get("Mobile_Normalized","") or "",
                        "Email":      row.get("Email_Normalized","")  or "",
                        "Website":    row.get("Website","")    or "",
                        "Location":   row.get("Location","")   or "",
                        "Notes":      row.get("Notes","")      or "",
                    }
                    if "Full Package (All Files)" in download_options:
                        write_full_package_to_zip(zipf, person, root=safe_root)
                    if "Business Cards Only" in download_options:
                        write_card_flat(zipf, person, root=safe_root)
                    if "Email Signatures Only" in download_options:
                        write_signature_flat(zipf, person, root=safe_root)

            zip_buf.seek(0)
            st.download_button("⬇️ Download Batch ZIP", zip_buf,
                               file_name=f"{safe_root}.zip",
                               key="batch_download_zip",
                               use_container_width=True)

# -------------------------------------------------
# FOOTER
# -------------------------------------------------
st.markdown(
    """
    <hr style="margin-top:3em; margin-bottom:0.5em; border: 0; border-top: 1px solid #e0e0e0;">
    <div style="text-align:center; color:gray; font-size:13px;">
        © Alraedah Finance — Internal Use Only
    </div>
    """,
    unsafe_allow_html=True
)

# -------------------------------------------------
# SIDEBAR (About & Info)
# -------------------------------------------------
with st.sidebar:
    st.image("assets/images/logo.png", use_container_width=True)
    st.markdown("### About Internal Hub")
    st.caption(
        "An internal tool developed for **Alraedah Finance** to automatically generate "
        "bilingual (English & Arabic) email signatures and business cards that align "
        "with the company’s brand identity."
    )
    st.divider()

    st.markdown("### How It Works")
    st.markdown(
        """
        1. **Single Entry** → Generate materials for one employee.  
        2. **Batch Upload** → Upload an Excel file with multiple employees.  
        3. **Download ZIP** → Get all signatures, cards, and QR files instantly.
        """
    )

    st.divider()

    st.markdown("### Output Files")
    st.markdown(
        """
        - Email Signature (EN / AR)  
        - Business Card (Front / Back)  
        - QR Code (PNG / SVG / VCF)  
        """
    )

    st.divider()
    st.caption("Version 1.0.1 · Built by Abdurrahman Alowain · For Internal Use Only")